import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px

# ── Настройка страницы ────────────────────────────────
st.set_page_config(
    page_title="Модель Леонтьева",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Стили ─────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
  html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
  .main { background: #f8f9fb; }

  /* Шапка */
  .app-header {
    background: linear-gradient(135deg, #0f2044 0%, #1a3a6e 100%);
    color: white; padding: 28px 36px; border-radius: 12px;
    margin-bottom: 24px;
  }
  .app-header h1 { font-size: 1.6rem; font-weight: 700; margin: 0; letter-spacing: -0.3px; }
  .app-header p  { font-size: 0.88rem; color: #a8c4e8; margin: 6px 0 0; }

  /* Карточки метрик */
  .metric-card {
    background: white; border-radius: 10px; padding: 18px 20px;
    box-shadow: 0 1px 8px rgba(0,0,0,0.07); border-left: 4px solid #1a3a6e;
  }
  .metric-label { font-size: 0.76rem; color: #7a8a9a; text-transform: uppercase; letter-spacing: 0.8px; }
  .metric-value { font-size: 1.5rem; font-weight: 700; color: #0f2044; margin-top: 4px; }
  .metric-sub   { font-size: 0.76rem; color: #a0aab4; margin-top: 2px; }

  /* Секции */
  .section-title {
    font-size: 0.78rem; font-weight: 700; text-transform: uppercase;
    letter-spacing: 1px; color: #1a3a6e; margin: 20px 0 10px;
    padding-bottom: 6px; border-bottom: 2px solid #e8edf5;
  }

  /* Шок-теги */
  .shock-tag {
    display: inline-block; padding: 3px 10px; border-radius: 20px;
    font-size: 0.78rem; font-weight: 600; margin: 2px;
  }
  .shock-pos { background: #d4edda; color: #155724; }
  .shock-neg { background: #f8d7da; color: #721c24; }

  /* Таблица */
  .stDataFrame { border-radius: 8px; overflow: hidden; }

  /* Sidebar */
  [data-testid="stSidebar"] {
    background: #0f2044 !important;
  }
  [data-testid="stSidebar"] * { color: #cfe0f5 !important; }
  [data-testid="stSidebar"] .stSelectbox label,
  [data-testid="stSidebar"] .stNumberInput label,
  [data-testid="stSidebar"] .stSlider label { color: #a8c4e8 !important; font-size: 0.82rem !important; }
  [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 { color: white !important; }

  div[data-testid="metric-container"] {
    background: white; border-radius: 10px; padding: 14px;
    box-shadow: 0 1px 6px rgba(0,0,0,0.06); border-left: 3px solid #1a3a6e;
  }
</style>
""", unsafe_allow_html=True)

# ── Шапка ─────────────────────────────────────────────
st.markdown("""
<div class="app-header">
  <h1>📊 Модель Леонтьева — Симулятор отраслевых шоков</h1>
  <p>Анализ межотраслевых связей и мультипликативных эффектов на основе таблиц «Затраты — Выпуск»</p>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════
# БОКОВАЯ ПАНЕЛЬ — загрузка и параметры
# ══════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## ⚙️ Параметры модели")
    st.markdown("---")

    uploaded = st.file_uploader("Загрузите файл Excel (ТЗВ)", type=["xlsx", "xls"])

    if uploaded:
        xl = pd.ExcelFile(uploaded)
        sheets = xl.sheet_names

        st.markdown("### 📋 Матрица L (лист 10)")
        sheet_L  = st.selectbox("Лист:", sheets, key="sl")
        z_r1     = st.number_input("Строка начала данных:", min_value=1, value=1, key="r1")
        z_r2     = st.number_input("Строка конца данных:", min_value=1, value=1, key="r2")
        z_c1     = st.number_input("Столбец начала матрицы:", min_value=1, value=1, key="c1")
        z_c2     = st.number_input("Столбец конца матрицы:", min_value=1, value=1, key="c2")
        names_c  = st.number_input("Столбец названий (0=нет):", min_value=0, value=0, key="nc")

        st.markdown("### 📈 ВДС и выпуск (лист 6)")
        sheet_V  = st.selectbox("Лист:", sheets, key="sv")
        vds_row  = st.number_input("Строка ВДС:", min_value=1, value=1, key="vr")
        x_row    = st.number_input("Строка выпуска:", min_value=1, value=1, key="xr")

        st.markdown("---")
        run = st.button("▶ Запустить модель", use_container_width=True)
    else:
        st.info("Загрузите файл Excel для начала работы")
        run = False

# ══════════════════════════════════════════════════════
# ЗАГРУЗКА ДАННЫХ
# ══════════════════════════════════════════════════════
@st.cache_data
def load_data(file_bytes, sheet_L, sheet_V, r1, r2, c1, c2, nc, vr, xr):
    import io
    from openpyxl import load_workbook
    import numpy as np

    # Читаем через openpyxl напрямую (надёжнее для сложных файлов)
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    def ws_to_df(sheet_name):
        ws = wb[sheet_name]
        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        return pd.DataFrame(data)

    df_L  = ws_to_df(sheet_L)
    df_vx = ws_to_df(sheet_V)
    wb.close()

    ri1 = r1-1; ri2 = r2
    ci1 = c1-1; ci2 = c2
    n   = ri2 - ri1

    names  = [str(df_L.iloc[ri1+i, nc-1]) for i in range(n)] if nc > 0 \
             else [f'Отрасль {i+1}' for i in range(n)]
    L      = df_L.iloc[ri1:ri2, ci1:ci2].values.astype(float)
    X_base = df_vx.iloc[xr-1, ci1:ci2].values.astype(float)
    vds    = df_vx.iloc[vr-1, ci1:ci2].values.astype(float)

    v     = np.where(X_base > 0, vds / X_base, 0)
    gdp_m = v @ L

    return names, L, X_base, vds, v, gdp_m, n

# ══════════════════════════════════════════════════════
# ОСНОВНОЙ ЭКРАН
# ══════════════════════════════════════════════════════
if uploaded and run:
    try:
        file_bytes = uploaded.read()
        names, L, X_base, vds, v, gdp_m, n = load_data(
            file_bytes, sheet_L, sheet_V,
            int(z_r1), int(z_r2), int(z_c1), int(z_c2), int(names_c),
            int(vds_row), int(x_row)
        )

        assert L.shape == (n, n), f"Матрица L не квадратная: {L.shape}"
        assert len(X_base) == n

        # Сохраняем в session_state
        st.session_state['model'] = {
            'names': names, 'L': L, 'X_base': X_base,
            'gdp_m': gdp_m, 'v': v, 'n': n
        }
        st.session_state['shocks'] = {}
        st.success(f"✓ Модель загружена: {n} отраслей")

    except Exception as e:
        st.error(f"Ошибка загрузки: {e}")

# ── Работа с моделью ──────────────────────────────────
if 'model' in st.session_state:
    md = st.session_state['model']
    names  = md['names']
    L      = md['L']
    X_base = md['X_base']
    gdp_m  = md['gdp_m']
    n      = md['n']

    if 'shocks' not in st.session_state:
        st.session_state['shocks'] = {}

    # ── Метрики верхнего уровня ───────────────────────
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Отраслей в модели", f"{n}")
    with col2:
        total_x = X_base.sum()
        st.metric("Суммарный выпуск", f"{total_x/1e12:.2f} трлн ₸")
    with col3:
        st.metric("Макс. мультипликатор ВВП", f"{gdp_m.max():.4f}")
    with col4:
        st.metric("Мин. мультипликатор ВВП", f"{gdp_m.min():.4f}")

    st.markdown("---")

    # ── Задание шоков ─────────────────────────────────
    st.markdown('<div class="section-title">Управление шоками конечного спроса</div>', unsafe_allow_html=True)

    col_l, col_r = st.columns([1, 2])

    with col_l:
        shock_mode = st.radio("Тип шока:", ["% от базового Y", "Млрд тенге (маржинальный)"], horizontal=True)
        sector_idx = st.selectbox("Выберите отрасль:", range(n), format_func=lambda i: names[i])

        if shock_mode == "% от базового Y":
            shock_val = st.slider("Изменение конечного спроса (%):", -100, 100, 0, 10)
            dY_abs = X_base[sector_idx] * shock_val / 100
            st.caption(f"Абсолютное изменение: {dY_abs/1e9:.2f} млрд ₸")
        else:
            shock_val = st.slider("Маржинальный шок (млрд тенге):", -100, 100, 0, 1)
            dY_abs = shock_val * 1e9
            st.caption(f"% от базы: {dY_abs/X_base[sector_idx]*100:.1f}%")

        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("+ Применить", use_container_width=True):
                if shock_val != 0:
                    st.session_state['shocks'][sector_idx] = {
                        'val': shock_val, 'mode': shock_mode, 'dY': dY_abs
                    }
        with col_b:
            if st.button("↺ Сбросить всё", use_container_width=True):
                st.session_state['shocks'] = {}

        # Список активных шоков
        if st.session_state['shocks']:
            st.markdown("**Активные шоки:**")
            for idx, sh in list(st.session_state['shocks'].items()):
                lbl = f"+{sh['val']}" if sh['val'] > 0 else str(sh['val'])
                unit = "%" if "%" in sh['mode'] else " млрд ₸"
                c1s, c2s = st.columns([4, 1])
                with c1s:
                    st.markdown(f"<small>{'🟢' if sh['val']>0 else '🔴'} <b>{names[idx][:40]}</b> {lbl}{unit}</small>", unsafe_allow_html=True)
                with c2s:
                    if st.button("✕", key=f"del_{idx}"):
                        del st.session_state['shocks'][idx]
                        st.rerun()
        else:
            st.info("Шоки не заданы — базовое состояние")

    # ── Расчёт ────────────────────────────────────────
    dY = np.zeros(n)
    for idx, sh in st.session_state['shocks'].items():
        dY[idx] = sh['dY']

    dX      = L @ dY
    X_plan  = X_base + dX
    dX_pct  = np.where(X_base > 0, dX / X_base * 100, 0)
    d_gdp   = gdp_m * dY

    with col_r:
        # ── Графики ───────────────────────────────────
        tab1, tab2, tab3, tab4 = st.tabs([
            "ΔX абсолютное", "ΔX % от базы",
            "Мультипликатор ВВП", "ΔВВП от шока"
        ])

        has_shock = bool(st.session_state['shocks'])

        def make_bar(values, title, fmt=".0f", color_scale=True):
            idx_sorted = np.argsort(np.abs(values))[::-1]
            if not has_shock:
                idx_sorted = np.argsort(X_base)[::-1][:20]
            show_n = min(n, len(idx_sorted))
            idx_show = idx_sorted[:show_n]

            colors = ['#27ae60' if v >= 0 else '#e74c3c' for v in values[idx_show]]
            short_names = [names[i][:30]+'…' if len(names[i])>30 else names[i] for i in idx_show]

            fig = go.Figure(go.Bar(
                x=values[idx_show],
                y=short_names,
                orientation='h',
                marker_color=colors,
                text=[f"{v:{fmt}}" for v in values[idx_show]],
                textposition='outside',
                hovertext=[names[i] for i in idx_show],
                hovertemplate='%{hovertext}<br>%{x}<extra></extra>'
            ))
            fig.update_layout(
                title=title, height=max(300, show_n * 22),
                margin=dict(l=10, r=60, t=40, b=10),
                plot_bgcolor='white', paper_bgcolor='white',
                font=dict(family='Inter', size=11),
                xaxis=dict(gridcolor='#f0f0f0', zerolinecolor='#ccc'),
                yaxis=dict(autorange='reversed')
            )
            return fig

        with tab1:
            fig = make_bar(dX, "Изменение валового выпуска ΔX (тенге)", ".2e")
            st.plotly_chart(fig, use_container_width=True)

        with tab2:
            fig = make_bar(dX_pct, "Изменение выпуска ΔX (% от базы)", ".2f")
            st.plotly_chart(fig, use_container_width=True)

        with tab3:
            idx_m = np.argsort(gdp_m)[::-1]
            short_names = [names[i][:30]+'…' if len(names[i])>30 else names[i] for i in idx_m]
            fig = go.Figure(go.Bar(
                x=gdp_m[idx_m], y=short_names, orientation='h',
                marker_color='#1a3a6e',
                text=[f"{v:.4f}" for v in gdp_m[idx_m]],
                textposition='outside',
                hovertext=[names[i] for i in idx_m],
                hovertemplate='%{hovertext}<br>Мульт. ВВП: %{x:.4f}<extra></extra>'
            ))
            fig.update_layout(
                title="Мультипликатор ВВП по отраслям (тенге ВВП / тенге спроса)",
                height=max(300, n * 22),
                margin=dict(l=10, r=80, t=40, b=10),
                plot_bgcolor='white', paper_bgcolor='white',
                font=dict(family='Inter', size=11),
                xaxis=dict(gridcolor='#f0f0f0'),
                yaxis=dict(autorange='reversed')
            )
            st.plotly_chart(fig, use_container_width=True)

        with tab4:
            fig = make_bar(d_gdp, "Изменение ВВП от шока ΔВВП (тенге)", ".2e")
            st.plotly_chart(fig, use_container_width=True)

    # ── Таблица результатов ───────────────────────────
    st.markdown('<div class="section-title">Полная таблица результатов</div>', unsafe_allow_html=True)

    df_result = pd.DataFrame({
        'Отрасль'         : names,
        'X база (млрд ₸)' : (X_base / 1e9).round(2),
        'X план (млрд ₸)' : (X_plan / 1e9).round(2),
        'ΔX (млрд ₸)'     : (dX / 1e9).round(2),
        'ΔX %'            : dX_pct.round(2),
        'Мульт. ВВП'      : gdp_m.round(4),
        'ΔВВП (млрд ₸)'   : (d_gdp / 1e9).round(2),
    })

    def color_delta(val):
        if isinstance(val, float):
            if val > 0: return 'color: #155724; font-weight: 600'
            if val < 0: return 'color: #721c24; font-weight: 600'
        return ''

    styled = df_result.style.applymap(
        color_delta, subset=['ΔX (млрд ₸)', 'ΔX %', 'ΔВВП (млрд ₸)']
    ).format({
        'X база (млрд ₸)': '{:,.2f}',
        'X план (млрд ₸)': '{:,.2f}',
        'ΔX (млрд ₸)'    : '{:+,.2f}',
        'ΔX %'           : '{:+.2f}%',
        'Мульт. ВВП'     : '{:.4f}',
        'ΔВВП (млрд ₸)'  : '{:+,.2f}',
    })

    st.dataframe(styled, use_container_width=True, height=400)

    # ── Экспорт ───────────────────────────────────────
    st.markdown('<div class="section-title">Экспорт результатов</div>', unsafe_allow_html=True)
    col_e1, col_e2 = st.columns(2)
    with col_e1:
        csv = df_result.to_csv(index=False, encoding='utf-8-sig')
        st.download_button("⬇ Скачать CSV", csv, "результаты_леонтьев.csv", "text/csv", use_container_width=True)
    with col_e2:
        import io
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df_result.to_excel(writer, sheet_name='Результаты', index=False)
            pd.DataFrame({'Мультипликатор ВВП': gdp_m}, index=names).to_excel(writer, sheet_name='Мультипликаторы')
        st.download_button("⬇ Скачать Excel", buf.getvalue(), "результаты_леонтьев.xlsx",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          use_container_width=True)

elif not uploaded:
    # ── Экран приветствия ─────────────────────────────
    st.markdown("""
    <div style="text-align:center;padding:60px 20px;color:#7a8a9a">
      <div style="font-size:4rem;margin-bottom:16px">📊</div>
      <h2 style="color:#0f2044;font-weight:700">Загрузите файл для начала работы</h2>
      <p style="font-size:1rem;max-width:500px;margin:0 auto">
        Загрузите файл Excel с таблицами «Затраты — Выпуск» через панель слева,
        укажите диапазоны данных и запустите модель.
      </p>
      <div style="margin-top:32px;display:flex;justify-content:center;gap:32px;flex-wrap:wrap">
        <div style="background:white;border-radius:10px;padding:20px 28px;box-shadow:0 2px 10px rgba(0,0,0,0.07);min-width:160px">
          <div style="font-size:1.8rem">🔢</div>
          <div style="font-weight:600;color:#0f2044;margin-top:8px">Матрица L</div>
          <div style="font-size:0.8rem;color:#7a8a9a">Коэф. полных затрат</div>
        </div>
        <div style="background:white;border-radius:10px;padding:20px 28px;box-shadow:0 2px 10px rgba(0,0,0,0.07);min-width:160px">
          <div style="font-size:1.8rem">📈</div>
          <div style="font-weight:600;color:#0f2044;margin-top:8px">Мультипликатор</div>
          <div style="font-size:0.8rem;color:#7a8a9a">ВВП по отраслям</div>
        </div>
        <div style="background:white;border-radius:10px;padding:20px 28px;box-shadow:0 2px 10px rgba(0,0,0,0.07);min-width:160px">
          <div style="font-size:1.8rem">⚡</div>
          <div style="font-weight:600;color:#0f2044;margin-top:8px">Симулятор шоков</div>
          <div style="font-size:0.8rem;color:#7a8a9a">% или млрд тенге</div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)
